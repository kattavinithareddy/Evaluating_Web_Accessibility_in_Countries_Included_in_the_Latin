from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime

def generate_pdf_report(project, report_data, report_type):
    """Generate PDF report"""
    filename = f"report_{project.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    doc = SimpleDocTemplate(filepath, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"<b>{report_type.replace('_', ' ').title()}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Project info
    project_info = Paragraph(f"<b>Project:</b> {project.project_name}<br/><b>URL:</b> {project.url}", styles['Normal'])
    elements.append(project_info)
    elements.append(Spacer(1, 20))
    
    # Compliance score
    score_text = Paragraph(f"<b>Compliance Score:</b> {report_data['project']['compliance_score']}%", styles['Heading2'])
    elements.append(score_text)
    elements.append(Spacer(1, 20))
    
    # Issues table
    issues_data = [
        ['Severity', 'Count'],
        ['Critical', report_data['issues']['critical']],
        ['Serious', report_data['issues']['serious']],
        ['Moderate', report_data['issues']['moderate']],
        ['Minor', report_data['issues']['minor']],
    ]
    
    table = Table(issues_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    return f'reports/{filename}'

def generate_excel_report(project, report_data, report_type):
    """Generate Excel report"""
    filename = f"report_{project.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Report"
    
    # Header
    ws['A1'] = report_type.replace('_', ' ').title()
    ws['A1'].font = Font(size=16, bold=True)
    
    # Project info
    ws['A3'] = "Project Name:"
    ws['B3'] = project.project_name
    ws['A4'] = "URL:"
    ws['B4'] = project.url
    ws['A5'] = "Compliance Score:"
    ws['B5'] = f"{report_data['project']['compliance_score']}%"
    
    # Issues
    ws['A7'] = "Issues Summary"
    ws['A7'].font = Font(size=14, bold=True)
    
    ws['A8'] = "Severity"
    ws['B8'] = "Count"
    ws['A9'] = "Critical"
    ws['B9'] = report_data['issues']['critical']
    ws['A10'] = "Serious"
    ws['B10'] = report_data['issues']['serious']
    ws['A11'] = "Moderate"
    ws['B11'] = report_data['issues']['moderate']
    ws['A12'] = "Minor"
    ws['B12'] = report_data['issues']['minor']
    
    wb.save(filepath)
    return f'reports/{filename}'

def generate_html_report(project, report_data, report_type):
    """Generate HTML report"""
    filename = f"report_{project.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>{report_type.replace('_', ' ').title()}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            h1 {{ color: #2c3e50; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th, td {{ border: 1px solid #bdc3c7; padding: 12px; text-align: left; }}
            th {{ background-color: #2c3e50; color: #ecf0f1; }}
        </style>
    </head>
    <body>
        <h1>{report_type.replace('_', ' ').title()}</h1>
        <p><strong>Project:</strong> {project.project_name}</p>
        <p><strong>URL:</strong> {project.url}</p>
        <p><strong>Compliance Score:</strong> {report_data['project']['compliance_score']}%</p>
        
        <h2>Issues Summary</h2>
        <table>
            <tr>
                <th>Severity</th>
                <th>Count</th>
            </tr>
            <tr>
                <td>Critical</td>
                <td>{report_data['issues']['critical']}</td>
            </tr>
            <tr>
                <td>Serious</td>
                <td>{report_data['issues']['serious']}</td>
            </tr>
            <tr>
                <td>Moderate</td>
                <td>{report_data['issues']['moderate']}</td>
            </tr>
            <tr>
                <td>Minor</td>
                <td>{report_data['issues']['minor']}</td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    with open(filepath, 'w') as f:
        f.write(html_content)
    
    return f'reports/{filename}'
